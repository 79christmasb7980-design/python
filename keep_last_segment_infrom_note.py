#!/usr/bin/env python3
"""
Keep only the last segment of text in column A of an Excel file where segments are
separated by patterns like '<number>'. Example cell content:

  "1>첫번째 내용 2>두번째 내용 3>마지막 내용"

This script will keep only "3>마지막 내용" for that cell.

Usage:
  python keep_last_segment_infrom_note.py [infrom_note.xlsx] [--inplace]

By default it reads `infrom_note.xlsx`, makes a timestamped backup, and writes the
modified workbook back to the same filename. Use --inplace to overwrite without creating
an additional backup (not recommended).
"""
import argparse
import os
import re
import shutil
import sys
from datetime import datetime
import unicodedata
from openpyxl.utils import get_column_letter

try:
    import openpyxl
except Exception as e:
    print("openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    raise


PATTERN = re.compile(r"\d+>")


def process_cell_value(value: str) -> str:
    """Return the substring starting from the last occurrence of a digit+">" pattern.

    If no pattern is found, return the original value.
    """
    if value is None:
        return value
    s = str(value)
    matches = list(PATTERN.finditer(s))
    if not matches:
        return s

    # If there are at least two occurrences, keep from the start of the second-last one.
    # Otherwise (only one), keep from the last occurrence.
    if len(matches) >= 2:
        start_idx = matches[-2].start()
    else:
        start_idx = matches[-1].start()

    # Keep from chosen start index to the end; strip leading/trailing whitespace
    result = s[start_idx:].strip()
    # Remove trailing date/datetime if present at the end of the string (various formats)
    return remove_trailing_date_time(result)


def remove_trailing_date_time(s: str) -> str:
    """Remove trailing date or datetime at the end of the string.

    Matches common forms like:
      - 2025-11-14
      - 2025/11/14
      - 2025.11.14
      - 20251114
      - 2025-11-14 12:34 or 12:34:56
      - 12:34 or 12:34:56
      - 2025년11월14일

    If such a pattern appears at the very end (optionally preceded by a separator), it will be removed.
    """
    if not s:
        return s

    # Patterns to detect trailing dates/times
    date_time_re = re.compile(
        r"(?:\s*[-–—]?\s*)?"
        #r"(?:(?:\d{4}[./-]\d{1,2}[./-]\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?)"  # 2025-11-14 or with time
        #am & pm 이 나오는 패턴은 제외
        r"(?:\d{4}[./-]\d{1,2}[./-]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?(?:\s?[APap][Mm])?)?)"  # 2025-11-14 or with time
        # r"|(?:\d{8})"  # 20251114
        # r"|(?:\d{1,2}:\d{2}(?::\d{2})?)"  # 12:34 or 12:34:56
        # r"|(?:\d{4}년\d{1,2}월\d{1,2}일(?:\s*\d{1,2}:\d{2})?))\s*$"
        , re.UNICODE,
    )

    new = re.sub(date_time_re, "", s).strip()
    return new


def display_length(s: str) -> int:
    """Estimate display length of a string counting East Asian wide/fullwidth chars as 2.

    Uses unicodedata.east_asian_width to approximate visual width so we can set column width
    in Excel more reasonably for mixed Latin/Korean text.
    """
    if not s:
        return 0
    total = 0
    for ch in s:
        try:
            ea = unicodedata.east_asian_width(ch)
        except Exception:
            ea = 'N'
        if ea in ('F', 'W'):
            total += 2
        else:
            total += 1
    return total
    max_width_per_line = 40

    row = 1
    while True:
        cell = ws[f"A{row}"]
        value = cell.value

        # 빈셀이면 loop 종료
        if value is None or str(value).strip() == "":
            break

        # 표시 길이 계산
        disp_len = display_length(str(value))

        # 필요 줄 수 계산
        lines = math.ceil(disp_len / max_width_per_line)

        # 행 전체 높이를 A열 텍스트 기준으로 조절
        ws.row_dimensions[row].height = 15 * lines

        row += 1




def process_workbook(path: str, sheet_name: str | None = None, inplace: bool = False) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    # Backup original unless user explicitly asks not to
    backup_path = None
    if not inplace:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{os.path.splitext(path)[0]}_backup_{timestamp}.xlsx"
        shutil.copy2(path, backup_path)

    wb = openpyxl.load_workbook(path)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Iterate rows and process column A (1) from first row to last used row
    max_row = ws.max_row
    changed = 0
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=1)
        original = cell.value
        if original is None:
            continue
        new = process_cell_value(original)
        if new != original:
            cell.value = new
            changed += 1

    # Auto-adjust all used columns' widths based on estimated display length
    try:
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            max_display_len = 0
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is None:
                    continue
                disp_len = display_length(str(val))
                if disp_len > max_display_len:
                    max_display_len = disp_len

            # Set column width with small padding and sensible caps
            col_letter = get_column_letter(col)
            col_width = max(8, min(max_display_len + 2, 120))
            ws.column_dimensions[col_letter].width = col_width
    except Exception:
        # Ignore width-setting errors and proceed to save
        pass

    wb.save(path)

    return backup_path if backup_path else path


def main():
    parser = argparse.ArgumentParser(description="Keep only the last '\\d+>' segment in column A of an Excel file.")
    parser.add_argument("file", nargs="?", default="infrom_note.xlsx", help="Excel file to process (default: infrom_note.xlsx)")
    parser.add_argument("--sheet", help="Sheet name to process (default: active sheet)")
    parser.add_argument("--inplace", action="store_true", help="Overwrite without making a backup")
    args = parser.parse_args() 

    try:
        backup = process_workbook(args.file, sheet_name=args.sheet, inplace=args.inplace)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        #sys.exit()                                                          
        sys.exit()
    if args.inplace:
        print(f"Processed and saved: {args.file} (inplace) ")
    else:
        print(f"Processed and saved: {args.file}")
        print(f"Backup of original created at: {backup}")


if __name__ == "__main__":
    main()
